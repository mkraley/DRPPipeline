"""
Compare claimed-but-not-uploaded URLs between the current Data_Inventories sheet
and a historical snapshot (either a Google Drive revision or a separate sheet copy).

Definition used:
- claimed but not uploaded = URL non-empty, at least one claimed column non-empty,
  and every Download Location column empty.

Usage:
    python scripts/tally_data_inventories/compare_claimed_unuploaded_history.py
    python scripts/tally_data_inventories/compare_claimed_unuploaded_history.py --days-ago 60
    python scripts/tally_data_inventories/compare_claimed_unuploaded_history.py --revision-id <id>
    python scripts/tally_data_inventories/compare_claimed_unuploaded_history.py --historical-sheet-id <sheet_id>
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import quote
import re

import requests
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2 import service_account
from openpyxl import load_workbook

_REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_REPO_ROOT))

from utils.Args import Args  # noqa: E402
from scripts.tally_data_inventories.sheet_claimed_tally import (  # noqa: E402
    _all_claimed_empty,
    _all_dl_empty,
    _cell_nonempty,
    _row_is_data_for_skips,
    find_claimed_columns_with_header_skips,
    find_download_location_columns_with_header_skips,
    find_named_column_index_and_skip,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_creds(credentials_path: Path):
    return service_account.Credentials.from_service_account_file(
        str(credentials_path),
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly",
        ],
    )


def _read_service_account_email(credentials_path: Path) -> str:
    try:
        raw = json.loads(credentials_path.read_text(encoding="utf-8"))
    except Exception:
        return ""
    return str(raw.get("client_email", "")).strip()


def _fetch_current_xlsx(sheet_id: str, token: str) -> bytes:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=300)
    if resp.status_code == 403:
        # Fallback: allow public/shared-link snapshots to be fetched without Bearer auth.
        resp = requests.get(url, timeout=300)
    resp.raise_for_status()
    data = resp.content
    if len(data) < 4 or data[:2] != b"PK":
        raise ValueError("Current export did not return XLSX content.")
    return data


def _extract_sheet_id(value: str) -> str:
    """Accept raw sheet id or full Google Sheets URL; return sheet id."""
    s = (value or "").strip()
    if not s:
        return ""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", s)
    if m:
        return m.group(1)
    return s


def _list_revisions(file_id: str, token: str) -> List[dict]:
    revisions: List[dict] = []
    page_token: Optional[str] = None
    while True:
        url = (
            f"https://www.googleapis.com/drive/v3/files/{quote(file_id, safe='')}/revisions"
            "?pageSize=200"
            "&fields=nextPageToken,revisions(id,modifiedTime,keepForever,exportLinks)"
            "&supportsAllDrives=true"
        )
        if page_token:
            url = f"{url}&pageToken={quote(page_token, safe='')}"
        resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=120)
        if resp.status_code == 403:
            raise PermissionError(
                "Drive revisions API returned 403. The service account may have Sheets read "
                "access but not sufficient Drive access to list revision history."
            )
        resp.raise_for_status()
        payload = resp.json()
        revisions.extend(payload.get("revisions", []))
        page_token = payload.get("nextPageToken")
        if not page_token:
            break
    return revisions


def _parse_google_ts(ts: str) -> dt.datetime:
    return dt.datetime.fromisoformat(ts.replace("Z", "+00:00"))


def _pick_revision(revisions: List[dict], days_ago: int, explicit_revision_id: Optional[str]) -> dict:
    if not revisions:
        raise ValueError("No revisions found for this file.")
    if explicit_revision_id:
        for rev in revisions:
            if str(rev.get("id")) == str(explicit_revision_id):
                return rev
        raise ValueError(f"Revision id {explicit_revision_id!r} not found.")

    now = dt.datetime.now(dt.timezone.utc)
    target = now - dt.timedelta(days=days_ago)
    with_ts: List[Tuple[dt.datetime, dict]] = []
    for rev in revisions:
        ts = rev.get("modifiedTime")
        if not ts:
            continue
        with_ts.append((_parse_google_ts(ts), rev))
    if not with_ts:
        raise ValueError("Revisions did not include modifiedTime.")

    with_ts.sort(key=lambda x: x[0])
    not_after = [item for item in with_ts if item[0] <= target]
    if not_after:
        return not_after[-1][1]
    return with_ts[0][1]


def _download_revision_xlsx(revision: dict, token: str) -> bytes:
    links = revision.get("exportLinks") or {}
    export_url = links.get(XLSX_MIME)
    if not export_url:
        raise ValueError(
            "Selected revision does not expose XLSX exportLinks; "
            "Google may not allow exporting this historical revision."
        )
    resp = requests.get(export_url, headers={"Authorization": f"Bearer {token}"}, timeout=300)
    resp.raise_for_status()
    data = resp.content
    if len(data) < 4 or data[:2] != b"PK":
        raise ValueError("Historical revision export did not return XLSX content.")
    return data


def _claimed_not_uploaded_urls(xlsx: bytes, url_column_name: str) -> Set[str]:
    out: Set[str] = set()
    wb = load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            r1 = next(it, None)
            if r1 is None:
                continue
            row1 = list(r1)
            r2 = next(it, None)
            row2 = list(r2) if r2 is not None else None

            claimed_cols, claimed_skips = find_claimed_columns_with_header_skips(row1, row2)
            dl_cols, dl_skips = find_download_location_columns_with_header_skips(row1, row2)
            url_match = find_named_column_index_and_skip(row1, row2, url_column_name)

            if not claimed_cols or not url_match:
                continue

            url_ci, url_skip = url_match
            combined = sorted(set(claimed_cols) | set(dl_cols))
            combined_skips = {**claimed_skips, **dl_skips}

            def maybe_add(row_list: List[object], sheet_row: int) -> None:
                if sheet_row <= url_skip:
                    return
                if not _row_is_data_for_skips(sheet_row, claimed_cols, claimed_skips):
                    return
                if combined and not _row_is_data_for_skips(sheet_row, combined, combined_skips):
                    return
                if not _cell_nonempty(row_list, url_ci):
                    return
                if _all_claimed_empty(row_list, claimed_cols):
                    return
                if not _all_dl_empty(row_list, dl_cols):
                    return
                out.add(str(row_list[url_ci]).strip())

            if row2 is not None:
                maybe_add(row2, 2)
                sheet_row = 3
            else:
                sheet_row = 2
            for row in it:
                maybe_add(list(row), sheet_row)
                sheet_row += 1
    finally:
        wb.close()
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare claimed-not-uploaded URLs with old revision")
    parser.add_argument("--days-ago", type=int, default=60, help="Target age in days (default: 60)")
    parser.add_argument(
        "--config",
        default="config.json",
        help="Path to config file (default: config.json)",
    )
    parser.add_argument(
        "--revision-id",
        default="",
        help="Optional explicit Google Drive revision id to compare instead of --days-ago",
    )
    parser.add_argument(
        "--historical-sheet-id",
        default="",
        help="Optional historical snapshot spreadsheet id (or full URL). If set, uses this sheet instead of Drive revisions.",
    )
    args = parser.parse_args()

    Args.initialize_from_config(Path(args.config))
    sheet_id = (getattr(Args, "google_sheet_id", None) or "").strip()
    creds_raw = getattr(Args, "google_credentials", None)
    url_col = (getattr(Args, "sourcing_url_column", None) or "URL").strip() or "URL"

    if not sheet_id or not creds_raw:
        print("Set google_sheet_id and google_credentials in config.", file=sys.stderr)
        raise SystemExit(1)

    creds_path = Path(creds_raw)
    if not creds_path.is_file():
        print(f"Credentials file not found: {creds_path}", file=sys.stderr)
        raise SystemExit(1)

    creds = _build_creds(creds_path)
    creds.refresh(GoogleAuthRequest())
    token = str(creds.token)

    service_account_email = _read_service_account_email(creds_path)
    try:
        current_xlsx = _fetch_current_xlsx(sheet_id, token)
    except requests.HTTPError as exc:
        print(f"Could not read current sheet as XLSX: {exc}", file=sys.stderr)
        if service_account_email:
            print(
                f"Share current sheet with service account: {service_account_email}",
                file=sys.stderr,
            )
        raise SystemExit(1)
    historical_sheet_id = _extract_sheet_id(args.historical_sheet_id)
    selected_id = ""
    selected_ts = ""
    if historical_sheet_id:
        try:
            historical_xlsx = _fetch_current_xlsx(historical_sheet_id, token)
        except requests.HTTPError as exc:
            print(f"Could not read historical sheet as XLSX: {exc}", file=sys.stderr)
            if service_account_email:
                print(
                    f'Share historical sheet "{historical_sheet_id}" with service account: '
                    f"{service_account_email}",
                    file=sys.stderr,
                )
            print(
                "Also make sure the sheet allows export/download.",
                file=sys.stderr,
            )
            raise SystemExit(1)
        selected_id = historical_sheet_id
        selected_ts = "n/a (separate historical spreadsheet copy)"
    else:
        try:
            revisions = _list_revisions(sheet_id, token)
        except PermissionError as exc:
            print(str(exc), file=sys.stderr)
            print(
                "Either share Drive revision access for this sheet, or pass "
                "--historical-sheet-id with a copied snapshot sheet.",
                file=sys.stderr,
            )
            raise SystemExit(1)
        selected = _pick_revision(revisions, args.days_ago, args.revision_id.strip() or None)
        selected_id = str(selected.get("id", ""))
        selected_ts = str(selected.get("modifiedTime", ""))
        historical_xlsx = _download_revision_xlsx(selected, token)

    current_urls = _claimed_not_uploaded_urls(current_xlsx, url_col)
    historical_urls = _claimed_not_uploaded_urls(historical_xlsx, url_col)

    overlap = current_urls & historical_urls
    current_only = current_urls - historical_urls

    print(f"Using URL column: {url_col}")
    if historical_sheet_id:
        print(f"Historical sheet id: {selected_id}")
        print(f"Historical sheet source: {selected_ts}")
    else:
        print(f"Historical revision id: {selected_id}")
        print(f"Historical revision modifiedTime: {selected_ts}")
    print()
    print(f"Current claimed-but-not-uploaded URLs: {len(current_urls)}")
    print(f"Historical claimed-but-not-uploaded URLs: {len(historical_urls)}")
    print(f"Current URLs that were also claimed-but-not-uploaded then: {len(overlap)}")
    if current_urls:
        pct = (len(overlap) / len(current_urls)) * 100
        print(f"Share of current claimed-not-uploaded also present then: {pct:.1f}%")
    print()
    print("Sample URLs currently claimed-not-uploaded but not in historical snapshot:")
    for url in sorted(current_only)[:25]:
        print(f"  {url}")


if __name__ == "__main__":
    main()
