"""
Create an Excel report comparing current vs historical "claimed but not uploaded" rows.

Outputs an XLSX workbook with:
1) overlap_rows: rows whose URL is claimed-but-not-uploaded in both snapshots.
   Columns: tab, claimant, title, source_url
2) current_claimant_tally: claimant counts for current claimed-but-not-uploaded rows.

Usage:
  python scripts/tally_data_inventories/export_claimed_unuploaded_overlap_report.py --historical-sheet "<id or URL>"
  python scripts/tally_data_inventories/export_claimed_unuploaded_overlap_report.py --historical-sheet "<id or URL>" --output claimed_overlap.xlsx
  python scripts/tally_data_inventories/export_claimed_unuploaded_overlap_report.py --historical-xlsx "C:\\path\\to\\snapshot.xlsx"
"""
from __future__ import annotations

import argparse
import io
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2 import service_account
from openpyxl import Workbook, load_workbook

_REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_REPO_ROOT))

from utils.Args import Args  # noqa: E402
from scripts.tally_data_inventories.sheet_claimed_tally import (  # noqa: E402
    _all_claimed_empty,
    _all_dl_empty,
    _cell_nonempty,
    _first_nonempty_claimed_value,
    _row_is_data_for_skips,
    find_claimed_columns_with_header_skips,
    find_download_location_columns_with_header_skips,
    find_named_column_index_and_skip,
)


@dataclass
class ClaimedRow:
    tab: str
    claimant: str
    title: str
    source_url: str


def _extract_sheet_id(value: str) -> str:
    s = (value or "").strip()
    if not s:
        return ""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", s)
    if m:
        return m.group(1)
    return s


def _read_service_account_email(credentials_path: Path) -> str:
    try:
        raw = json.loads(credentials_path.read_text(encoding="utf-8"))
    except Exception:
        return ""
    return str(raw.get("client_email", "")).strip()


def _build_token(credentials_path: Path) -> str:
    creds = service_account.Credentials.from_service_account_file(
        str(credentials_path),
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    creds.refresh(GoogleAuthRequest())
    return str(creds.token)


def _fetch_sheet_xlsx(sheet_id: str, token: str) -> bytes:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=300)
    if resp.status_code in (401, 403):
        # fallback for publicly shared sheets
        resp = requests.get(url, timeout=300)
    resp.raise_for_status()
    data = resp.content
    if len(data) < 4 or data[:2] != b"PK":
        raise ValueError("Export did not return XLSX.")
    return data


def _find_title_column_index(row1: List[object], row2: Optional[List[object]]) -> Optional[int]:
    # Prefer exact "title", then common variants containing "title".
    exact = find_named_column_index_and_skip(row1, row2, "title")
    if exact is not None:
        return exact[0]
    len1 = len(row1)
    len2 = len(row2) if row2 else 0
    max_cols = max(len1, len2)
    for i in range(max_cols):
        v1 = str(row1[i] if i < len1 else "").strip().lower()
        v2 = str(row2[i] if row2 and i < len2 else "").strip().lower()
        merged = f"{v1} {v2}".strip()
        if "title" in merged:
            return i
    return None


def _extract_claimed_rows(xlsx: bytes, url_column_name: str) -> Dict[str, ClaimedRow]:
    rows_by_url: Dict[str, ClaimedRow] = {}
    wb = load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            r1 = next(it, None)
            if r1 is None:
                continue
            row1 = list(r1)
            r2 = next(it, None)
            row2 = list(r2) if r2 is not None else None

            claimed_cols, claimed_skips = find_claimed_columns_with_header_skips(row1, row2)
            if not claimed_cols:
                continue
            dl_cols, dl_skips = find_download_location_columns_with_header_skips(row1, row2)
            url_match = find_named_column_index_and_skip(row1, row2, url_column_name)
            if url_match is None:
                continue

            title_ci = _find_title_column_index(row1, row2)
            url_ci, url_skip = url_match
            combined = sorted(set(claimed_cols) | set(dl_cols))
            combined_skips = {**claimed_skips, **dl_skips}

            def handle_row(row_list: List[object], sheet_row: int) -> None:
                if sheet_row <= url_skip:
                    return
                if not _row_is_data_for_skips(sheet_row, claimed_cols, claimed_skips):
                    return
                if combined and not _row_is_data_for_skips(sheet_row, combined, combined_skips):
                    return
                if not _cell_nonempty(row_list, url_ci):
                    return
                if _all_claimed_empty(row_list, claimed_cols):
                    return
                if not _all_dl_empty(row_list, dl_cols):
                    return

                source_url = str(row_list[url_ci]).strip()
                claimant = _first_nonempty_claimed_value(row_list, claimed_cols) or ""
                title = ""
                if title_ci is not None and title_ci < len(row_list):
                    title = str(row_list[title_ci] or "").strip()
                if source_url and source_url not in rows_by_url:
                    rows_by_url[source_url] = ClaimedRow(
                        tab=ws.title,
                        claimant=claimant,
                        title=title,
                        source_url=source_url,
                    )

            if row2 is not None:
                handle_row(row2, 2)
                sheet_row = 3
            else:
                sheet_row = 2
            for row in it:
                handle_row(list(row), sheet_row)
                sheet_row += 1
    finally:
        wb.close()
    return rows_by_url


def _write_report(
    output_path: Path,
    overlap_rows: List[ClaimedRow],
    current_tally: List[Tuple[str, int]],
) -> None:
    wb = Workbook()
    ws_overlap = wb.active
    ws_overlap.title = "overlap_rows"
    ws_overlap.append(["tab", "claimant", "title", "source_url"])
    for r in overlap_rows:
        ws_overlap.append([r.tab, r.claimant, r.title, r.source_url])

    ws_tally = wb.create_sheet("current_claimant_tally")
    ws_tally.append(["claimant", "count"])
    for claimant, count in current_tally:
        ws_tally.append([claimant, count])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export overlap of claimed-not-uploaded rows")
    parser.add_argument(
        "--historical-sheet",
        default="",
        help="Historical spreadsheet id or full URL",
    )
    parser.add_argument(
        "--historical-xlsx",
        default="",
        help="Local historical XLSX path (alternative to --historical-sheet)",
    )
    parser.add_argument(
        "--output",
        default="claimed_unuploaded_overlap_report.xlsx",
        help="Output XLSX path (default: claimed_unuploaded_overlap_report.xlsx)",
    )
    parser.add_argument(
        "--config",
        default="config.json",
        help="Path to config file (default: config.json)",
    )
    args = parser.parse_args()

    Args.initialize_from_config(Path(args.config))
    sheet_id = (getattr(Args, "google_sheet_id", None) or "").strip()
    creds_raw = getattr(Args, "google_credentials", None)
    url_col = (getattr(Args, "sourcing_url_column", None) or "URL").strip() or "URL"
    hist_id = _extract_sheet_id(args.historical_sheet)
    hist_xlsx_path = Path(args.historical_xlsx).expanduser() if args.historical_xlsx else None

    if not sheet_id or not creds_raw:
        print("Missing google_sheet_id/google_credentials.", file=sys.stderr)
        raise SystemExit(1)
    if not hist_id and hist_xlsx_path is None:
        print("Provide either --historical-sheet or --historical-xlsx.", file=sys.stderr)
        raise SystemExit(1)

    creds_path = Path(creds_raw)
    if not creds_path.is_file():
        print(f"Credentials file not found: {creds_path}", file=sys.stderr)
        raise SystemExit(1)

    token = _build_token(creds_path)
    svc_email = _read_service_account_email(creds_path)

    try:
        current_xlsx = _fetch_sheet_xlsx(sheet_id, token)
    except Exception as exc:
        print(f"Unable to read current sheet ({sheet_id}): {exc}", file=sys.stderr)
        if svc_email:
            print(f"Share current sheet with service account: {svc_email}", file=sys.stderr)
        raise SystemExit(1)

    if hist_xlsx_path is not None:
        if not hist_xlsx_path.is_file():
            print(f"Historical XLSX not found: {hist_xlsx_path}", file=sys.stderr)
            raise SystemExit(1)
        historical_xlsx = hist_xlsx_path.read_bytes()
    else:
        try:
            historical_xlsx = _fetch_sheet_xlsx(hist_id, token)
        except Exception as exc:
            print(f"Unable to read historical sheet ({hist_id}): {exc}", file=sys.stderr)
            if svc_email:
                print(f"Share historical sheet with service account: {svc_email}", file=sys.stderr)
            raise SystemExit(1)

    current_rows = _extract_claimed_rows(current_xlsx, url_col)
    historical_rows = _extract_claimed_rows(historical_xlsx, url_col)

    overlap_urls = sorted(set(current_rows.keys()) & set(historical_rows.keys()))
    overlap = [current_rows[url] for url in overlap_urls]

    tally: Dict[str, int] = {}
    for r in current_rows.values():
        key = r.claimant or "(blank)"
        tally[key] = tally.get(key, 0) + 1
    tally_sorted = sorted(tally.items(), key=lambda x: (-x[1], x[0].lower()))

    out_path = Path(args.output)
    _write_report(out_path, overlap, tally_sorted)

    print(f"Output report: {out_path}")
    print(f"Current claimed-not-uploaded URLs: {len(current_rows)}")
    print(f"Historical claimed-not-uploaded URLs: {len(historical_rows)}")
    print(f"Overlap URLs: {len(overlap_urls)}")


if __name__ == "__main__":
    main()
