"""
Tally non-empty cells in columns whose header (row 1 and/or 2) includes the word \"claimed\".

Uses a single authenticated XLSX export (docs.google.com export URL) — not per-cell Sheets API
reads — then parses locally. Used by debug/tally_claimed_all_tabs.py.
"""

from __future__ import annotations

import collections
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Counter, Dict, List, Optional, Tuple

import requests
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2 import service_account
from openpyxl import load_workbook

# Whole word only — avoids matching \"unclaimed\", \"disclaimed\", etc.
_CLAIMED_WORD = re.compile(r"(?i)\bclaimed\b")

_DL_SUBSTRING = "download location"


@dataclass
class ClaimedTallyReport:
    """Result of scanning all worksheets for claimed columns and non-empty cells."""

    tally: Counter[str]
    """Counts per claimant name (trimmed cell text)."""

    sheets_without_claimed_column: tuple[str, ...]
    """Worksheets where no column header in row 1 or row 2 includes the word \"claimed\" (incl. empty tabs)."""

    sheets_without_download_location_column: tuple[str, ...]
    """Worksheets where no column header in row 1 or row 2 contains the substring \"download location\"."""

    sheets_missing_claimed_or_download_location: tuple[str, ...]
    """Union: tabs missing a claimed header, a download location header, or both."""

    unclaimed_url_rows_by_sheet: tuple[tuple[str, int], ...]
    """Per-tab counts: rows with URL filled (see ``url_column_name``) and all claimed columns empty.

    Only tabs that have at least one claimed column and a matching URL column header. Sorted by tab name.
    """

    sheets_without_url_column: tuple[str, ...]
    """Tabs that have a claimed column but no column in rows 1 or 2 whose header exactly matches ``url_column_name``."""

    url_column_name: Optional[str]
    """Column name used for URL detection, or None if unclaimed-URL tally was skipped."""

    claimed_without_download_location_by_claimant: tuple[tuple[str, int], ...]
    """Rows with at least one claimed cell filled and every Download Location column empty.

    Only tabs that have both claimed and Download Location headers. Sorted by count descending, then name.
    """

    claimed_without_download_location_by_sheet: tuple[tuple[str, int], ...]
    """Same rows as ``claimed_without_download_location_by_claimant``, counted by tab. Sorted by tab name."""

    @property
    def total_claimed_entries(self) -> int:
        """Total non-empty cells counted in all matching columns (sum of occurrences)."""
        return int(sum(self.tally.values()))

    @property
    def unique_claimant_count(self) -> int:
        """Number of distinct claimant strings."""
        return len(self.tally)


def header_cell_matches_claimed(text: object) -> bool:
    """True if the cell text includes the whole word \"claimed\" (case-insensitive)."""
    s = str(text or "").strip()
    return bool(_CLAIMED_WORD.search(s))


def header_cell_matches_download_location(text: object) -> bool:
    """True if the cell text contains the substring \"download location\" (case-insensitive)."""
    return _DL_SUBSTRING in str(text or "").strip().lower()


def find_claimed_columns_with_header_skips(
    row1: List[object],
    row2: Optional[List[object]],
) -> Tuple[List[int], Dict[int, int]]:
    """
    Columns where row 1 or row 2's cell includes the word \"claimed\".

    Returns:
        (column_indices, skip_rows_per_column). ``skip`` is the number of leading rows treated as
        header for that column (1 or 2). Data is counted only on rows with 1-based index > skip.
    """
    len1 = len(row1)
    len2 = len(row2) if row2 else 0
    max_cols = max(len1, len2)
    claimed_cols: List[int] = []
    skips: Dict[int, int] = {}
    for i in range(max_cols):
        v1 = row1[i] if i < len1 else None
        v2 = row2[i] if row2 and i < len2 else None
        c1 = header_cell_matches_claimed(v1)
        c2 = header_cell_matches_claimed(v2)
        if not (c1 or c2):
            continue
        if c1 and c2:
            skip = 2
        elif c2 and not c1:
            skip = 2
        else:
            skip = 1
        claimed_cols.append(i)
        skips[i] = skip
    return claimed_cols, skips


def find_download_location_columns_with_header_skips(
    row1: List[object],
    row2: Optional[List[object]],
) -> Tuple[List[int], Dict[int, int]]:
    """Columns where row 1 or row 2 header contains the substring \"download location\" (same skip rules as claimed)."""
    len1 = len(row1)
    len2 = len(row2) if row2 else 0
    max_cols = max(len1, len2)
    dl_cols: List[int] = []
    skips: Dict[int, int] = {}
    for i in range(max_cols):
        v1 = row1[i] if i < len1 else None
        v2 = row2[i] if row2 and i < len2 else None
        c1 = header_cell_matches_download_location(v1)
        c2 = header_cell_matches_download_location(v2)
        if not (c1 or c2):
            continue
        if c1 and c2:
            skip = 2
        elif c2 and not c1:
            skip = 2
        else:
            skip = 1
        dl_cols.append(i)
        skips[i] = skip
    return dl_cols, skips


def find_named_column_index_and_skip(
    row1: List[object],
    row2: Optional[List[object]],
    column_name: str,
) -> Optional[Tuple[int, int]]:
    """
    First column index where row 1 or row 2 header cell equals ``column_name`` (case-insensitive, stripped).

    Returns ``(column_index, skip_rows)`` using the same 1- vs 2-row header rule as claimed columns.
    """
    name_l = column_name.strip().lower()
    if not name_l:
        return None
    len1 = len(row1)
    len2 = len(row2) if row2 else 0
    max_cols = max(len1, len2)
    for i in range(max_cols):
        v1 = str(row1[i] if i < len1 else "").strip().lower()
        v2 = str(row2[i] if row2 and i < len2 else "").strip().lower()
        n1 = v1 == name_l
        n2 = v2 == name_l
        if not (n1 or n2):
            continue
        if n1 and n2:
            skip = 2
        elif n2 and not n1:
            skip = 2
        else:
            skip = 1
        return (i, skip)
    return None


def _cell_nonempty(row: List[object], ci: int) -> bool:
    if ci >= len(row):
        return False
    return bool(str(row[ci] or "").strip())


def _all_claimed_empty(row: List[object], claimed_cols: List[int]) -> bool:
    for ci in claimed_cols:
        if _cell_nonempty(row, ci):
            return False
    return True


def _all_dl_empty(row: List[object], dl_cols: List[int]) -> bool:
    for ci in dl_cols:
        if _cell_nonempty(row, ci):
            return False
    return True


def _first_nonempty_claimed_value(row: List[object], claimed_cols: List[int]) -> Optional[str]:
    for ci in claimed_cols:
        if _cell_nonempty(row, ci):
            return str(row[ci]).strip()
    return None


def _row_is_data_for_skips(sheet_row: int, col_indices: List[int], skips: Dict[int, int]) -> bool:
    for ci in col_indices:
        if sheet_row <= skips[ci]:
            return False
    return True


def _row_is_data_for_url_and_claimed(
    sheet_row: int,
    url_skip: int,
    claimed_cols: List[int],
    claimed_skips: Dict[int, int],
) -> bool:
    if sheet_row <= url_skip:
        return False
    for ci in claimed_cols:
        if sheet_row <= claimed_skips[ci]:
            return False
    return True


def _tally_cell(tally: Counter[str], row: List[object], ci: int) -> None:
    if ci >= len(row):
        return
    val = row[ci]
    if val is None:
        return
    cell = str(val).strip()
    if cell:
        tally[cell] += 1


def _maybe_count_unclaimed_url_row(
    sheet_row: int,
    row_list: List[object],
    url_ci: int,
    url_skip: int,
    claimed_cols: List[int],
    claimed_skips: Dict[int, int],
) -> bool:
    if not _row_is_data_for_url_and_claimed(sheet_row, url_skip, claimed_cols, claimed_skips):
        return False
    if not _cell_nonempty(row_list, url_ci):
        return False
    if not _all_claimed_empty(row_list, claimed_cols):
        return False
    return True


def _maybe_claimed_without_download_location(
    sheet_row: int,
    row_list: List[object],
    claimed_cols: List[int],
    claimed_skips: Dict[int, int],
    dl_cols: List[int],
    dl_skips: Dict[int, int],
) -> Optional[str]:
    if not dl_cols:
        return None
    combined_indices = sorted(set(claimed_cols) | set(dl_cols))
    combined_skips = {**claimed_skips, **dl_skips}
    if not _row_is_data_for_skips(sheet_row, combined_indices, combined_skips):
        return None
    claimant = _first_nonempty_claimed_value(row_list, claimed_cols)
    if not claimant:
        return None
    if not _all_dl_empty(row_list, dl_cols):
        return None
    return claimant


def tally_claimed_from_xlsx_bytes(
    data: bytes,
    url_column_name: Optional[str] = None,
) -> ClaimedTallyReport:
    """
    Scan every worksheet: inspect row 1 and row 2 for column headers containing the word
    \"claimed\", then tally non-empty cells below the per-column header depth.

    If ``url_column_name`` is set, also counts per sheet rows where that column is non-empty
    and every claimed column is empty (rows 1 and 2 header rules apply to the URL column too).
    """
    tally: Counter[str] = collections.Counter()
    no_claim: List[str] = []
    no_dl: List[str] = []
    no_url: List[str] = []
    unclaimed_by_sheet: Dict[str, int] = {}
    claimed_no_dl_by_claimant: Counter[str] = collections.Counter()
    claimed_no_dl_by_sheet: Dict[str, int] = {}

    url_name = (url_column_name or "").strip() or None

    bio = io.BytesIO(data)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            r1 = next(it, None)
            if r1 is None:
                no_claim.append(ws.title)
                no_dl.append(ws.title)
                continue
            row1 = list(r1)
            r2 = next(it, None)
            row2 = list(r2) if r2 is not None else None

            cols, skips = find_claimed_columns_with_header_skips(row1, row2)
            dl_cols, dl_skips = find_download_location_columns_with_header_skips(row1, row2)

            if not cols:
                no_claim.append(ws.title)
            if not dl_cols:
                no_dl.append(ws.title)

            if not cols:
                continue

            url_match: Optional[Tuple[int, int]] = None
            if url_name:
                url_match = find_named_column_index_and_skip(row1, row2, url_name)
                if url_match is None:
                    no_url.append(ws.title)

            unclaimed_delta = 0

            if row2 is not None:
                sheet_row = 2
                for ci in cols:
                    if sheet_row > skips[ci]:
                        _tally_cell(tally, row2, ci)
                if url_match is not None:
                    url_ci, url_skip = url_match
                    if _maybe_count_unclaimed_url_row(sheet_row, row2, url_ci, url_skip, cols, skips):
                        unclaimed_delta += 1
                cl = _maybe_claimed_without_download_location(
                    sheet_row, row2, cols, skips, dl_cols, dl_skips
                )
                if cl is not None:
                    claimed_no_dl_by_claimant[cl] += 1
                    claimed_no_dl_by_sheet[ws.title] = claimed_no_dl_by_sheet.get(ws.title, 0) + 1
                sheet_row = 3
                for row in it:
                    row_list = list(row)
                    for ci in cols:
                        if sheet_row > skips[ci]:
                            _tally_cell(tally, row_list, ci)
                    if url_match is not None:
                        url_ci, url_skip = url_match
                        if _maybe_count_unclaimed_url_row(
                            sheet_row, row_list, url_ci, url_skip, cols, skips
                        ):
                            unclaimed_delta += 1
                    cl = _maybe_claimed_without_download_location(
                        sheet_row, row_list, cols, skips, dl_cols, dl_skips
                    )
                    if cl is not None:
                        claimed_no_dl_by_claimant[cl] += 1
                        claimed_no_dl_by_sheet[ws.title] = (
                            claimed_no_dl_by_sheet.get(ws.title, 0) + 1
                        )
                    sheet_row += 1
            else:
                sheet_row = 2
                for row in it:
                    row_list = list(row)
                    for ci in cols:
                        if sheet_row > skips[ci]:
                            _tally_cell(tally, row_list, ci)
                    if url_match is not None:
                        url_ci, url_skip = url_match
                        if _maybe_count_unclaimed_url_row(
                            sheet_row, row_list, url_ci, url_skip, cols, skips
                        ):
                            unclaimed_delta += 1
                    cl = _maybe_claimed_without_download_location(
                        sheet_row, row_list, cols, skips, dl_cols, dl_skips
                    )
                    if cl is not None:
                        claimed_no_dl_by_claimant[cl] += 1
                        claimed_no_dl_by_sheet[ws.title] = (
                            claimed_no_dl_by_sheet.get(ws.title, 0) + 1
                        )
                    sheet_row += 1

            if url_match is not None:
                unclaimed_by_sheet[ws.title] = unclaimed_delta
    finally:
        wb.close()

    no_claim_sorted = tuple(sorted(no_claim))
    no_dl_sorted = tuple(sorted(no_dl))
    missing_either = tuple(sorted(set(no_claim) | set(no_dl)))
    no_url_sorted = tuple(sorted(no_url))
    unclaimed_tuple = tuple(sorted(unclaimed_by_sheet.items()))
    by_claimant_no_dl = tuple(
        sorted(claimed_no_dl_by_claimant.items(), key=lambda x: (-x[1], x[0].lower()))
    )
    by_sheet_no_dl = tuple(sorted(claimed_no_dl_by_sheet.items()))
    return ClaimedTallyReport(
        tally=tally,
        sheets_without_claimed_column=no_claim_sorted,
        sheets_without_download_location_column=no_dl_sorted,
        sheets_missing_claimed_or_download_location=missing_either,
        unclaimed_url_rows_by_sheet=unclaimed_tuple,
        sheets_without_url_column=no_url_sorted,
        url_column_name=url_name,
        claimed_without_download_location_by_claimant=by_claimant_no_dl,
        claimed_without_download_location_by_sheet=by_sheet_no_dl,
    )


def _fetch_spreadsheet_xlsx(spreadsheet_id: str, credentials_path: Path) -> bytes:
    """
    Download the entire spreadsheet as XLSX using a Bearer token (one HTTP request).

    Does not use spreadsheets.values.*; avoids Sheets API read quota for per-range calls.
    """
    creds = service_account.Credentials.from_service_account_file(
        str(credentials_path),
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    creds.refresh(GoogleAuthRequest())
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {creds.token}"},
        timeout=300,
    )
    resp.raise_for_status()
    data = resp.content
    if len(data) < 4 or data[:2] != b"PK":
        raise ValueError(
            "Export did not return an XLSX file (wrong credentials, no access, or invalid ID)."
        )
    return data


def tally_claimed_across_tabs(
    spreadsheet_id: str,
    credentials_path: Path,
    url_column_name: Optional[str] = None,
) -> ClaimedTallyReport:
    """
    For every worksheet, find columns whose header in row 1 or 2 includes the word \"claimed\",
    then count each non-empty data cell by value.

    One export request + local parsing (no Sheets values API batching per tab).
    """
    xlsx = _fetch_spreadsheet_xlsx(spreadsheet_id, credentials_path)
    return tally_claimed_from_xlsx_bytes(xlsx, url_column_name=url_column_name)
